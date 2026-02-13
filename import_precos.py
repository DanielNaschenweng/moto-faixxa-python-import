#!/usr/bin/env python3
"""
Script para importar dados da planilha de preços de peças de moto para MongoDB.
"""

import sys
from datetime import datetime
from openpyxl import load_workbook
from pymongo import MongoClient, ReplaceOne

# Configurações
MONGO_URI = "mongodb+srv://plxuser:6zCFmIAMlDe8ZEj0@plx-mongodb.yx2fpuj.mongodb.net/moto_faixxa?retryWrites=true&w=majority&appName=plx-mongodb&authSource=admin"
DATABASE_NAME = "moto_faixxa"
COLLECTION_NAME = "precos"
COLLECTION_NUVEMSHOP = "produtos_nuvemshop"
EXCEL_PATH = "/home/daniel/projetos_sh/moto_faixxa/TABELA PREÇOS NOTE 09-2023(2).xlsx"
IMAGES_PATH = "/home/daniel/projetos_sh/moto_faixxa/000 FAIXAS LIMPAS"


def conectar_mongodb():
    """Conecta ao MongoDB e retorna a coleção."""
    client = MongoClient(MONGO_URI)
    db = client[DATABASE_NAME]
    return db[COLLECTION_NAME], client


def processar_ano_modelo(modelo):
    """
    Processa o modelo para extrair e converter ano de 2 para 4 dígitos.
    Ex: "HAYABUSA 08" -> ("HAYABUSA 2008", 2008, "HAYABUSA 08")
        "CBR 1000 98" -> ("CBR 1000 1998", 1998, "CBR 1000 98")
        "GSX 750" -> ("GSX 750", None, None)

    Retorna: (modelo_formatado, ano, modelo_original_2dig)
    """
    if not modelo:
        return modelo, None, None

    import re
    partes = modelo.strip().split()

    if not partes:
        return modelo, None, None

    ultima_parte = partes[-1]

    # Verificar se a última parte é um número de 2 dígitos
    if re.match(r'^\d{2}$', ultima_parte):
        modelo_original = modelo.strip()
        ano_2dig = int(ultima_parte)
        # Converter para 4 dígitos: >= 50 é 19xx, < 50 é 20xx
        if ano_2dig >= 50:
            ano_4dig = 1900 + ano_2dig
        else:
            ano_4dig = 2000 + ano_2dig

        # Substituir última parte pelo ano com 4 dígitos
        partes[-1] = str(ano_4dig)
        modelo_formatado = " ".join(partes)
        return modelo_formatado, ano_4dig, modelo_original

    # Verificar se já tem 4 dígitos (19xx ou 20xx)
    if re.match(r'^(19|20)\d{2}$', ultima_parte):
        # Gerar versão com 2 dígitos para busca de compatibilidade
        ano_4dig = int(ultima_parte)
        ano_2dig = ano_4dig % 100
        partes_2dig = partes[:-1] + [f"{ano_2dig:02d}"]
        modelo_2dig = " ".join(partes_2dig)
        return modelo, ano_4dig, modelo_2dig

    # Não tem ano
    return modelo, None, None


def gerar_handle(texto):
    """Gera um handle (slug) a partir do texto."""
    import re
    import unicodedata
    # Normalizar unicode e remover acentos
    texto = unicodedata.normalize('NFKD', texto)
    texto = texto.encode('ASCII', 'ignore').decode('ASCII')
    # Converter para minúsculas e substituir espaços/caracteres especiais por hífen
    texto = texto.lower()
    texto = re.sub(r'[^a-z0-9]+', '-', texto)
    texto = re.sub(r'-+', '-', texto)  # Remover hífens duplicados
    return texto.strip('-')


def normalizar_texto(texto):
    """Normaliza texto para comparação (remove acentos, uppercase, espaços extras)."""
    import unicodedata
    if not texto:
        return ""
    texto = unicodedata.normalize('NFKD', str(texto))
    texto = texto.encode('ASCII', 'ignore').decode('ASCII')
    texto = texto.upper().strip()
    # Normalizar separadores: / vira espaço
    texto = texto.replace('/', ' ')
    # Remover espaços duplicados
    texto = ' '.join(texto.split())
    return texto


def indexar_imagens(caminho_base):
    """Percorre a pasta de imagens recursivamente e retorna um índice estruturado.

    Retorna dict: {marca: [{'path': caminho, 'path_norm': normalizado, 'imagens': [lista]}]}
    """
    import os
    indice = {}

    if not os.path.exists(caminho_base):
        print(f"AVISO: Pasta de imagens não encontrada: {caminho_base}")
        return indice

    def tem_imagens(caminho):
        """Verifica se a pasta contém imagens."""
        for arquivo in os.listdir(caminho):
            if arquivo.lower().endswith(('.jpg', '.jpeg', '.png')):
                return True
        return False

    def listar_imagens(caminho):
        """Lista imagens de uma pasta."""
        imagens = []
        for arquivo in os.listdir(caminho):
            if arquivo.lower().endswith(('.jpg', '.jpeg', '.png')):
                imagens.append(arquivo)
        return sorted(imagens)

    def percorrer_recursivo(caminho, marca, path_partes):
        """Percorre recursivamente as pastas."""
        for item in os.listdir(caminho):
            caminho_item = os.path.join(caminho, item)
            if not os.path.isdir(caminho_item):
                continue

            novas_partes = path_partes + [item]

            # Verificar se esta pasta tem imagens
            if tem_imagens(caminho_item):
                imagens = listar_imagens(caminho_item)
                if imagens:
                    # Construir path relativo e normalizado
                    path_relativo = "/" + "/".join(novas_partes)
                    # Normalizar o caminho completo (sem a marca)
                    path_sem_marca = " ".join(novas_partes[1:])  # Exclui a marca
                    path_norm = normalizar_texto(path_sem_marca)

                    indice[marca].append({
                        'path': path_relativo,
                        'path_norm': path_norm,
                        'imagens': imagens
                    })

            # Continuar recursão em subpastas
            percorrer_recursivo(caminho_item, marca, novas_partes)

    # Percorrer marcas (primeiro nível)
    for marca in os.listdir(caminho_base):
        caminho_marca = os.path.join(caminho_base, marca)
        if not os.path.isdir(caminho_marca):
            continue

        marca_norm = normalizar_texto(marca)
        if marca_norm not in indice:
            indice[marca_norm] = []

        percorrer_recursivo(caminho_marca, marca_norm, [marca])

    return indice


def buscar_imagens_produto(marca, modelo, cor, indice_imagens):
    """Busca imagens correspondentes a um produto.

    Retorna lista de dicts: [{'filename': str, 'position': int, 'path': str}]
    """
    marca_norm = normalizar_texto(marca)
    modelo_norm = normalizar_texto(modelo)
    cor_norm = normalizar_texto(cor)

    if marca_norm not in indice_imagens:
        return []

    pastas_marca = indice_imagens[marca_norm]

    # Palavras do modelo são OBRIGATÓRIAS
    palavras_modelo = [p for p in modelo_norm.split() if len(p) > 1]
    palavras_cor = [p for p in cor_norm.split() if len(p) > 2] if cor_norm else []

    # Estratégia de busca: encontrar pasta que contenha TODAS as palavras do modelo
    melhor_match = None
    melhor_score = 0

    for dados in pastas_marca:
        path_norm = dados['path_norm']

        # Verificar se TODAS as palavras do modelo estão no path
        modelo_matches = 0
        for palavra in palavras_modelo:
            if palavra in path_norm:
                modelo_matches += 1

        # Se não encontrou todas as palavras do modelo, pular
        if modelo_matches < len(palavras_modelo):
            continue

        # Calcular score baseado nas palavras de cor encontradas
        score = modelo_matches * 10  # Base: palavras do modelo

        cor_matches = 0
        for palavra in palavras_cor:
            if palavra in path_norm:
                cor_matches += 1
                score += 3

        # Bonus se encontrou todas as cores
        if palavras_cor and cor_matches == len(palavras_cor):
            score += 5

        if score > melhor_score:
            melhor_score = score
            melhor_match = dados

    if not melhor_match:
        return []

    # Filtrar e ordenar imagens
    arquivos = melhor_match['imagens']

    # 1. Filtrar: se existe "X PRIMEIRA.jpg", remover "X.jpg"
    arquivos_filtrados = []
    for arquivo in arquivos:
        nome_base = arquivo.rsplit('.', 1)[0]  # Remove extensão
        # Verificar se existe versão "PRIMEIRA" deste arquivo
        versao_primeira = f"{nome_base} PRIMEIRA"
        tem_primeira = any(
            a.rsplit('.', 1)[0] == versao_primeira
            for a in arquivos
        )
        # Se não tem versão PRIMEIRA, ou se este É a versão PRIMEIRA, manter
        if not tem_primeira or "PRIMEIRA" in nome_base:
            arquivos_filtrados.append(arquivo)

    # 2. Ordenar: priorizar imagens com marca/modelo no nome
    def prioridade_imagem(filename):
        nome_upper = filename.upper()
        score = 0
        # Verificar se contém palavras do modelo
        for palavra in palavras_modelo:
            if palavra in nome_upper:
                score -= 10  # Negativo para ficar primeiro na ordenação
        # Verificar se contém marca
        if marca_norm in nome_upper:
            score -= 5
        # KIT geralmente é a imagem principal
        if "KIT" in nome_upper:
            score -= 3
        return (score, filename)

    arquivos_ordenados = sorted(arquivos_filtrados, key=prioridade_imagem)

    # Montar lista de imagens
    imagens = []
    for i, filename in enumerate(arquivos_ordenados, 1):
        imagens.append({
            'filename': filename,
            'position': i,
            'path': melhor_match['path']
        })

    return imagens


def gerar_image_id(path, filename):
    """Gera um hash único para a imagem."""
    import hashlib
    texto = f"{path}/{filename}"
    return hashlib.md5(texto.encode()).hexdigest()


def encontrar_imagem_variante(peca, images):
    """Encontra a imagem correspondente à variante baseado no nome da peça.

    Retorna o id da imagem ou None se não encontrar.
    """
    if not peca or not images:
        return None

    peca_norm = normalizar_texto(peca)
    palavras_peca = peca_norm.split()

    # Mapeamento de sinônimos/abreviações
    mapeamento = {
        "ESQ": ["ESQUERDA", "ESQ"],
        "DIR": ["DIREITA", "DIR"],
        "RABET": ["RABETA", "RABET"],
        "TNQ": ["TANQUE", "TNQ"],
    }

    melhor_match = None
    melhor_score = 0

    for img in images:
        filename_norm = normalizar_texto(img['filename'])
        score = 0

        # Verificar cada palavra da peça
        for palavra in palavras_peca:
            if len(palavra) < 2:
                continue

            # Verificar match direto
            if palavra in filename_norm:
                score += 10
                continue

            # Verificar sinônimos
            for abrev, sinonimos in mapeamento.items():
                if palavra == abrev or palavra in sinonimos:
                    for sin in sinonimos:
                        if sin in filename_norm:
                            score += 10
                            break

        if score > melhor_score:
            melhor_score = score
            melhor_match = img

    # Retornar apenas se teve match significativo
    if melhor_match and melhor_score >= 10:
        return melhor_match.get('id')

    return None


TEXTO_DESCRICAO_PADRAO = """<p>Nossas faixas adesivas são confeccionadas em processo exatamente igual das originais.</p>
<p>Os materiais adesivos utilizados são <strong>Oracal 651</strong>. Estes materiais são os melhores disponíveis no mercado brasileiro e estão entre os melhores do mundo, o que garante tranquilidade ao consumidor em ter em mãos um produto que se aproxima bastante ao produto original em termos de qualidade.</p>
<p>O processo de impressão é o <strong>serigrafico</strong> para possibilitar fidelidade e solidez das cores, isto é, não desbotam e tem as tonalidades o mais próximo possível das originais. Dizemos isso pois ainda que na maioria dos casos seja possível reproduzir com fidelidade as cores das originais, em alguns tentamos aproximar ao máximo pois dependemos de pigmentos especiais que não existem no mercado brasileiro. No momento da compra questione sobre isso, caso não seja alertado por nós, para que você saiba exatamente o que está comprando.</p>"""


def gerar_descricao_produto(marca, modelo, cor, pecas):
    """Gera descrição completa do produto com dados e texto padrão em HTML."""
    # Formatar lista de peças disponíveis
    if pecas:
        formato_venda = ", ".join(pecas)
    else:
        formato_venda = "Peças avulsas"

    # Montar descrição com HTML
    descricao = f"""<h2>Faixa Adesiva para {marca} {modelo}</h2>

<h3>Destaques do Produto</h3>
<ul>
<li><strong>Material:</strong> Adesivo vinílico</li>
<li><strong>Formato de venda:</strong> {formato_venda}</li>
<li><strong>Cor:</strong> {cor if cor else "Consulte"}</li>
<li><strong>À prova d'água</strong></li>
</ul>

<hr>

<h3>Nosso Produto</h3>
{TEXTO_DESCRICAO_PADRAO}"""

    return descricao


def converter_para_nuvemshop(documentos, indice_imagens=None):
    """Converte documentos do formato interno para formato Nuvemshop."""
    produtos_nuvemshop = []
    produtos_com_imagem = 0
    total_imagens = 0

    for doc in documentos:
        marca = doc.get("marca", "")
        modelo = doc.get("modelo", "")
        modelo_antigo = doc.get("modelo_antigo")
        cor = doc.get("cor", "")
        ano = doc.get("ano")

        # Buscar imagens do produto
        images = []
        if indice_imagens:
            images = buscar_imagens_produto(marca, modelo, cor, indice_imagens)
            if images:
                produtos_com_imagem += 1
                total_imagens += len(images)

        # Adicionar id hash a cada imagem
        for img in images:
            img['id'] = gerar_image_id(img['path'], img['filename'])

        # Nome do produto: Marca + Modelo
        nome_produto = f"{marca} {modelo}".strip()

        # Handle: slug do nome + cor
        handle_base = f"{nome_produto} {cor}".strip() if cor else nome_produto
        handle = gerar_handle(handle_base)

        # Handle antigo (para busca de compatibilidade no upsert)
        handle_antigo = None
        if modelo_antigo:
            nome_produto_antigo = f"{marca} {modelo_antigo}".strip()
            handle_base_antigo = f"{nome_produto_antigo} {cor}".strip() if cor else nome_produto_antigo
            handle_antigo = gerar_handle(handle_base_antigo)

        # Converter variantes
        variantes_nuvemshop = []
        for i, var in enumerate(doc.get("variantes", [])):
            peca = var.get("peca") or ""
            elemento = var.get("elemento")

            # Values para identificar a variante (usando a peça como valor)
            values = [{"pt": peca}]
            if elemento:
                values.append({"pt": elemento})

            # Encontrar imagem correspondente à variante
            image_id = encontrar_imagem_variante(peca, images)

            variante_ns = {
                "position": i + 1,
                "sku": var.get("referencia"),
                "price": var.get("preco", 0.0),
                "stock": 1,
                "stock_management": True,
                "values": values,
                "imageId": image_id,
            }
            variantes_nuvemshop.append(variante_ns)

        # Extrair lista de peças para descrição
        pecas = [var.get("peca") for var in doc.get("variantes", []) if var.get("peca")]

        # Montar produto Nuvemshop
        produto_ns = {
            "name": {"pt": nome_produto},
            "description": {"pt": gerar_descricao_produto(marca, modelo, cor, pecas)},
            "handle": handle,
            "handle_antigo": handle_antigo,  # Para busca de compatibilidade no upsert
            "published": True,
            "requires_shipping": True,
            "height": 2.0,      # cm
            "width": 40.0,      # cm
            "depth": 50.0,      # cm (comprimento)
            "weight": 0.1,      # kg (100g)
            "attributes": [
                {"pt": "Peça"}
            ],
            "variants": variantes_nuvemshop,
            "images": images,
            "marca": marca,
            "modelo": modelo,
            "cor": cor,
            "ano": ano,
            "created_at": datetime.now(),
            "updated_at": datetime.now(),
        }
        produtos_nuvemshop.append(produto_ns)

    print(f"  Produtos com imagens: {produtos_com_imagem}/{len(documentos)}")
    print(f"  Total de imagens encontradas: {total_imagens}")

    return produtos_nuvemshop


def eh_linha_separadora(row):
    """Verifica se é linha separadora (tudo vazio exceto preço=0)."""
    modelo = row[0]
    cor = row[1]
    kit = row[2]
    preco = row[4]
    loc = row[5]
    ref = row[6]

    # Tudo vazio exceto preço que pode ser 0
    sem_modelo = not modelo or not str(modelo).strip()
    sem_cor = not cor or not str(cor).strip()
    sem_kit = not kit or not str(kit).strip()
    sem_loc = not loc or not str(loc).strip()
    sem_ref = not ref or not str(ref).strip()
    preco_zero = preco == 0 or preco is None

    return sem_modelo and sem_cor and sem_kit and sem_loc and sem_ref and preco_zero


def eh_linha_marca(row):
    """Verifica se é linha de marca (só primeira coluna preenchida, resto vazio).

    Marcas verdadeiras são curtas e não contêm números ou 'GAV'.
    Ex: AGRALE, APRILIA, BMW, HARLEY DAVIDSON, KASINSKI
    """
    modelo = row[0]
    cor = row[1]
    kit = row[2]
    elem = row[3]
    preco = row[4]
    loc = row[5]
    ref = row[6]

    tem_modelo = modelo and str(modelo).strip()
    sem_cor = not cor or not str(cor).strip()
    sem_kit = not kit or not str(kit).strip()
    sem_elem = not elem or not str(elem).strip()
    sem_preco = not preco or preco == 0
    sem_loc = not loc or not str(loc).strip()
    sem_ref = not ref or not str(ref).strip()

    if not (tem_modelo and sem_cor and sem_kit and sem_elem and sem_preco and sem_loc and sem_ref):
        return False

    # Marcas verdadeiras não contêm números nem "GAV"
    texto = str(modelo).strip().upper()
    tem_numero = any(c.isdigit() for c in texto)
    tem_gav = "GAV" in texto

    return not tem_numero and not tem_gav


def processar_planilha(caminho_arquivo, debug=False, debug_modelo=None):
    """Lê a planilha e retorna lista de documentos para inserção."""
    wb = load_workbook(caminho_arquivo, data_only=True)
    documentos = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        marca_aba = sheet_name.upper()
        eh_aba_outras = (marca_aba == "OUTRAS")

        print(f"Processando aba: {marca_aba} ({ws.max_row} linhas)")

        # Primeiro passo: agrupar linhas por bloco (separados por linha vazia com preço=0)
        # Na aba OUTRAS, também identificar linhas de marca
        blocos = []
        bloco_atual = []
        marca_atual = marca_aba  # Para abas normais, usa o nome da aba

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Na aba OUTRAS, verificar se é linha de marca
            if eh_aba_outras and eh_linha_marca(row):
                # Salvar bloco anterior se existir
                if bloco_atual:
                    blocos.append({"marca": marca_atual, "linhas": bloco_atual})
                    bloco_atual = []
                # Atualizar marca atual
                marca_atual = str(row[0]).strip().upper()
            elif eh_linha_separadora(row):
                if bloco_atual:
                    blocos.append({"marca": marca_atual, "linhas": bloco_atual})
                    bloco_atual = []
            else:
                bloco_atual.append(row)

        if bloco_atual:
            blocos.append({"marca": marca_atual, "linhas": bloco_atual})

        # Segundo passo: processar cada bloco
        for bloco in blocos:
            marca = bloco["marca"]
            linhas = bloco["linhas"]

            # Encontrar primeiro modelo e concatenar todas as cores
            modelo = None
            cores = []

            for linha in linhas:
                if not modelo and linha[0] and str(linha[0]).strip():
                    modelo = str(linha[0]).strip()
                # Coletar todas as cores (não duplicadas)
                if linha[1] and str(linha[1]).strip():
                    cor_linha = str(linha[1]).strip()
                    if cor_linha not in cores:
                        cores.append(cor_linha)

            # Concatenar cores na ordem em que aparecem
            cor = " ".join(cores) if cores else None

            # Processar ano do modelo (converter 2 dígitos para 4)
            modelo, ano, modelo_antigo = processar_ano_modelo(modelo)

            # Debug: mostrar detalhes do bloco
            if debug and modelo and (not debug_modelo or debug_modelo.upper() in modelo.upper()):
                print(f"\n[DEBUG] Bloco encontrado: {marca} - {modelo}")
                print(f"[DEBUG] Cores encontradas: {cores}")
                print(f"[DEBUG] Cor final: {cor}")
                print(f"[DEBUG] Linhas do bloco:")
                for i, linha in enumerate(linhas):
                    print(f"  [{i}] col0={repr(linha[0])} | col1={repr(linha[1])} | col2={repr(linha[2])}")

            # Coletar todas as variantes do bloco
            variantes = []
            for linha in linhas:
                kit_conjunto = linha[2]
                elem = linha[3]
                preco = linha[4]
                loc = linha[5]
                referencia = linha[6]

                if kit_conjunto or (preco and preco != 0):
                    variante = {
                        "peca": str(kit_conjunto).strip() if kit_conjunto else None,
                        "elemento": str(elem).strip() if elem else None,
                        "preco": float(preco) if preco and isinstance(preco, (int, float)) else 0.0,
                        "localizacao": str(loc).strip() if loc else None,
                        "referencia": str(referencia).strip() if referencia else None,
                    }
                    variantes.append(variante)

            # Criar documento único com todas as variantes
            if variantes:
                doc = {
                    "marca": marca,
                    "modelo": modelo,
                    "modelo_antigo": modelo_antigo,  # Para busca de compatibilidade no upsert
                    "cor": cor,
                    "ano": ano,
                    "variantes": variantes,
                    "data_importacao": datetime.now()
                }
                documentos.append(doc)

    wb.close()

    # Consolidar documentos com mesma chave (marca, modelo, cor)
    # mesclando suas variantes
    docs_consolidados = {}
    for doc in documentos:
        chave = (doc["marca"], doc["modelo"], doc["cor"])
        if chave in docs_consolidados:
            # Mesclar variantes
            docs_consolidados[chave]["variantes"].extend(doc["variantes"])
        else:
            docs_consolidados[chave] = doc

    documentos_finais = list(docs_consolidados.values())
    if len(documentos_finais) < len(documentos):
        print(f"  Consolidados {len(documentos)} blocos em {len(documentos_finais)} documentos")

    return documentos_finais


def remover_duplicados(db, colecao_nome, campo_chave):
    """Remove documentos duplicados mantendo apenas o mais recente."""
    colecao = db[colecao_nome]

    # Encontrar handles duplicados
    pipeline = [
        {"$group": {"_id": f"${campo_chave}", "count": {"$sum": 1}, "ids": {"$push": "$_id"}}},
        {"$match": {"count": {"$gt": 1}}}
    ]

    duplicados = list(colecao.aggregate(pipeline))
    if duplicados:
        print(f"  Encontrados {len(duplicados)} {campo_chave}s duplicados, removendo...")
        total_removidos = 0
        for dup in duplicados:
            # Manter o primeiro (mais antigo), remover os demais
            ids_para_remover = dup["ids"][1:]
            colecao.delete_many({"_id": {"$in": ids_para_remover}})
            total_removidos += len(ids_para_remover)
        print(f"  Removidos {total_removidos} documentos duplicados")


def remover_duplicados_compostos(db, colecao_nome, campos):
    """Remove documentos duplicados por chave composta mantendo apenas o mais recente."""
    colecao = db[colecao_nome]

    # Construir agrupamento por campos compostos
    group_id = {campo: f"${campo}" for campo in campos}

    pipeline = [
        {"$group": {"_id": group_id, "count": {"$sum": 1}, "ids": {"$push": "$_id"}}},
        {"$match": {"count": {"$gt": 1}}}
    ]

    duplicados = list(colecao.aggregate(pipeline))
    if duplicados:
        print(f"  Encontrados {len(duplicados)} registros duplicados em {colecao_nome}, removendo...")
        total_removidos = 0
        for dup in duplicados:
            ids_para_remover = dup["ids"][1:]
            colecao.delete_many({"_id": {"$in": ids_para_remover}})
            total_removidos += len(ids_para_remover)
        print(f"  Removidos {total_removidos} documentos duplicados")


def criar_indices(db):
    """Cria índices para otimizar as operações de upsert."""
    # Remover duplicados antes de criar índices únicos
    remover_duplicados_compostos(db, COLLECTION_NAME, ["marca", "modelo", "cor"])
    remover_duplicados(db, COLLECTION_NUVEMSHOP, "handle")

    # Dropar índices antigos se existirem (para recriar com unique=True)
    try:
        db[COLLECTION_NAME].drop_index("idx_marca_modelo_cor")
    except Exception:
        pass  # Índice não existe, ok

    try:
        db[COLLECTION_NUVEMSHOP].drop_index("idx_handle")
    except Exception:
        pass  # Índice não existe, ok

    # Índice composto único para coleção precos
    db[COLLECTION_NAME].create_index(
        [("marca", 1), ("modelo", 1), ("cor", 1)],
        name="idx_marca_modelo_cor",
        unique=True,
        background=True
    )

    # Índice único para coleção produtos_nuvemshop (handle é a chave de busca)
    db[COLLECTION_NUVEMSHOP].create_index(
        "handle",
        name="idx_handle",
        unique=True,
        background=True
    )


def main():
    """Função principal."""
    # Verificar argumentos
    modo_clean = "--clean" in sys.argv
    modo_debug = "--debug" in sys.argv

    # Modelo para debug (ex: --debug HAYABUSA)
    debug_modelo = None
    for arg in sys.argv:
        if arg.startswith("--debug="):
            modo_debug = True
            debug_modelo = arg.split("=", 1)[1]
            break
    if modo_debug and not debug_modelo:
        # Se --debug sem modelo, mostra todos
        debug_modelo = ""

    print("=" * 60)
    print("Importador de Preços para MongoDB")
    print(f"Modo: {'CLEAN (limpar base)' if modo_clean else 'UPSERT (manter IDs)'}")
    print("=" * 60)

    # Conectar ao MongoDB
    print("\nConectando ao MongoDB...")
    try:
        colecao, client = conectar_mongodb()
        db = client[DATABASE_NAME]
        print(f"Conectado: {MONGO_URI}")
        print(f"Database: {DATABASE_NAME}")
        print(f"Collection: {COLLECTION_NAME}")

        # Criar índices (operação idempotente)
        print("Verificando índices...")
        criar_indices(db)
    except Exception as e:
        print(f"Erro ao conectar ao MongoDB: {e}")
        return

    # Processar planilha
    print(f"\nLendo planilha: {EXCEL_PATH}")
    documentos = processar_planilha(EXCEL_PATH, debug=modo_debug, debug_modelo=debug_modelo)
    print(f"\nTotal de documentos para inserir: {len(documentos)}")

    if not documentos:
        print("Nenhum documento para inserir.")
        client.close()
        return

    # Inserir/Atualizar documentos
    count_antes = colecao.count_documents({})

    if modo_clean:
        # Modo CLEAN: limpar e inserir tudo
        if count_antes > 0:
            print(f"\nRemovendo {count_antes} documentos existentes...")
            colecao.delete_many({})

        print("\nInserindo documentos...")
        try:
            resultado = colecao.insert_many(documentos)
            print(f"Documentos inseridos: {len(resultado.inserted_ids)}")
        except Exception as e:
            print(f"Erro ao inserir documentos: {e}")
            client.close()
            return
    else:
        # Modo UPSERT: atualizar existentes, inserir novos (mantém IDs)
        print(f"\nAtualizando documentos (existentes: {count_antes})...")
        try:
            # Buscar todos os documentos existentes para criar mapa de lookup
            docs_existentes = {}
            for doc_existente in colecao.find({}, {"_id": 1, "marca": 1, "modelo": 1, "cor": 1}):
                chave = (doc_existente["marca"], doc_existente.get("modelo"), doc_existente.get("cor"))
                docs_existentes[chave] = doc_existente["_id"]

            # Preparar operações em lote
            operacoes = []
            for doc in documentos:
                doc["data_importacao"] = datetime.now()

                # Tentar encontrar documento existente (pelo modelo novo ou antigo)
                chave_nova = (doc["marca"], doc["modelo"], doc["cor"])
                chave_antiga = (doc["marca"], doc.get("modelo_antigo"), doc["cor"]) if doc.get("modelo_antigo") else None

                doc_id = docs_existentes.get(chave_nova)
                if not doc_id and chave_antiga:
                    doc_id = docs_existentes.get(chave_antiga)

                if doc_id:
                    # Documento existe - fazer replace pelo _id
                    filtro = {"_id": doc_id}
                else:
                    # Documento novo - usar filtro por campos
                    filtro = {
                        "marca": doc["marca"],
                        "modelo": doc["modelo"],
                        "cor": doc["cor"]
                    }

                # Remover modelo_antigo do documento final (não precisa salvar)
                doc_para_salvar = {k: v for k, v in doc.items() if k != "modelo_antigo"}
                operacoes.append(ReplaceOne(filtro, doc_para_salvar, upsert=True))

            # Executar em lote
            resultado = colecao.bulk_write(operacoes, ordered=False)
            print(f"Documentos inseridos: {resultado.upserted_count}")
            print(f"Documentos atualizados: {resultado.modified_count}")
        except Exception as e:
            print(f"Erro ao atualizar documentos: {e}")
            client.close()
            return

    # Verificar inserção
    count_depois = colecao.count_documents({})
    print(f"\nTotal de documentos na coleção: {count_depois}")

    # Mostrar exemplo
    print("\nExemplo de documento inserido:")
    exemplo = colecao.find_one()
    print(f"  marca: {exemplo.get('marca')}")
    print(f"  modelo: {exemplo.get('modelo')}")
    print(f"  cor: {exemplo.get('cor')}")
    print(f"  variantes: {len(exemplo.get('variantes', []))} itens")
    for i, var in enumerate(exemplo.get('variantes', [])[:3]):
        print(f"    [{i}] {var.get('peca')} - R$ {var.get('preco')}")
    if len(exemplo.get('variantes', [])) > 3:
        print(f"    ... e mais {len(exemplo.get('variantes', [])) - 3} variantes")
    print(f"  data_importacao: {exemplo.get('data_importacao')}")

    # Estatísticas por marca
    print("\nDocumentos por marca:")
    pipeline = [{"$group": {"_id": "$marca", "total": {"$sum": 1}}}]
    for item in colecao.aggregate(pipeline):
        print(f"  {item['_id']}: {item['total']}")

    # ========================================
    # Coleção Nuvemshop
    # ========================================
    print("\n" + "=" * 60)
    print("Gerando produtos no formato Nuvemshop...")
    print("=" * 60)

    # Indexar imagens
    print("\nIndexando imagens...")
    indice_imagens = indexar_imagens(IMAGES_PATH)
    total_pastas = sum(len(lista) for lista in indice_imagens.values())
    print(f"  Marcas encontradas: {len(indice_imagens)}")
    print(f"  Pastas com imagens: {total_pastas}")

    # Converter para formato Nuvemshop
    print("\nConvertendo produtos...")
    produtos_nuvemshop = converter_para_nuvemshop(documentos, indice_imagens)
    print(f"\nTotal de produtos Nuvemshop: {len(produtos_nuvemshop)}")

    # Obter coleção Nuvemshop
    db = client[DATABASE_NAME]
    colecao_nuvemshop = db[COLLECTION_NUVEMSHOP]

    # Inserir/Atualizar produtos Nuvemshop
    count_ns_antes = colecao_nuvemshop.count_documents({})

    if modo_clean:
        # Modo CLEAN: limpar e inserir tudo
        if count_ns_antes > 0:
            print(f"Removendo {count_ns_antes} documentos existentes...")
            colecao_nuvemshop.delete_many({})

        print("Inserindo produtos Nuvemshop...")
        try:
            resultado_ns = colecao_nuvemshop.insert_many(produtos_nuvemshop)
            print(f"Produtos inseridos: {len(resultado_ns.inserted_ids)}")
        except Exception as e:
            print(f"Erro ao inserir produtos Nuvemshop: {e}")
            client.close()
            return
    else:
        # Modo UPSERT: atualizar existentes, inserir novos (mantém IDs)
        print(f"Atualizando produtos Nuvemshop (existentes: {count_ns_antes})...")
        try:
            # Buscar todos os documentos existentes para criar mapa de lookup
            docs_existentes_ns = {}
            for doc_existente in colecao_nuvemshop.find({}, {"_id": 1, "handle": 1, "created_at": 1}):
                docs_existentes_ns[doc_existente["handle"]] = {
                    "_id": doc_existente["_id"],
                    "created_at": doc_existente.get("created_at")
                }

            # Preparar operações em lote
            operacoes = []
            for produto in produtos_nuvemshop:
                produto["updated_at"] = datetime.now()

                # Tentar encontrar documento existente (pelo handle novo ou antigo)
                doc_existente = docs_existentes_ns.get(produto["handle"])
                if not doc_existente and produto.get("handle_antigo"):
                    doc_existente = docs_existentes_ns.get(produto["handle_antigo"])

                # Manter created_at original se existir
                if doc_existente and doc_existente.get("created_at"):
                    produto["created_at"] = doc_existente["created_at"]

                if doc_existente:
                    # Documento existe - fazer replace pelo _id
                    filtro = {"_id": doc_existente["_id"]}
                else:
                    # Documento novo - usar filtro por handle
                    filtro = {"handle": produto["handle"]}

                # Remover handle_antigo do documento final (não precisa salvar)
                produto_para_salvar = {k: v for k, v in produto.items() if k != "handle_antigo"}
                operacoes.append(ReplaceOne(filtro, produto_para_salvar, upsert=True))

            # Executar em lote
            resultado = colecao_nuvemshop.bulk_write(operacoes, ordered=False)
            print(f"Produtos inseridos: {resultado.upserted_count}")
            print(f"Produtos atualizados: {resultado.modified_count}")
        except Exception as e:
            print(f"Erro ao atualizar produtos Nuvemshop: {e}")
            client.close()
            return

    # Verificar inserção
    count_ns_depois = colecao_nuvemshop.count_documents({})
    print(f"Total na coleção {COLLECTION_NUVEMSHOP}: {count_ns_depois}")

    # Mostrar exemplo Nuvemshop
    print("\nExemplo de produto Nuvemshop:")
    exemplo_ns = colecao_nuvemshop.find_one()
    print(f"  name: {exemplo_ns.get('name')}")
    print(f"  handle: {exemplo_ns.get('handle')}")
    print(f"  variants: {len(exemplo_ns.get('variants', []))} itens")
    for i, var in enumerate(exemplo_ns.get('variants', [])[:2]):
        print(f"    [{i}] sku: {var.get('sku')} - R$ {var.get('price')}")
    if len(exemplo_ns.get('variants', [])) > 2:
        print(f"    ... e mais {len(exemplo_ns.get('variants', [])) - 2} variantes")
    print(f"  images: {len(exemplo_ns.get('images', []))} imagens")
    for i, img in enumerate(exemplo_ns.get('images', [])[:3]):
        print(f"    [{i}] {img.get('filename')} (path: {img.get('path')})")
    if len(exemplo_ns.get('images', [])) > 3:
        print(f"    ... e mais {len(exemplo_ns.get('images', [])) - 3} imagens")

    # Buscar um produto com imagens para mostrar
    exemplo_com_img = colecao_nuvemshop.find_one({"images.0": {"$exists": True}})
    if exemplo_com_img and exemplo_com_img != exemplo_ns:
        print("\nExemplo de produto COM imagens:")
        print(f"  name: {exemplo_com_img.get('name')}")
        print(f"  images: {len(exemplo_com_img.get('images', []))} imagens")
        for i, img in enumerate(exemplo_com_img.get('images', [])[:3]):
            print(f"    [{i}] {img.get('filename')} (path: {img.get('path')})")

    client.close()
    print("\n" + "=" * 60)
    print("Importação concluída com sucesso!")
    print(f"  - {COLLECTION_NAME}: {count_depois} documentos")
    print(f"  - {COLLECTION_NUVEMSHOP}: {count_ns_depois} produtos")
    print("=" * 60)


if __name__ == "__main__":
    main()
